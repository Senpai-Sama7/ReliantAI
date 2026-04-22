"""
Dataset file inspection helpers.
"""
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import json
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook  # type: ignore[import-untyped]


@dataclass(frozen=True)
class DatasetInspection:
    """Normalized dataset metadata extracted from an uploaded file."""

    row_count: int
    column_count: int
    columns: list[str]
    file_type: str


def build_upload_path(upload_root: Path, dataset_id: str, original_filename: str) -> Path:
    """Build a stable upload path for a stored dataset file."""
    suffix = Path(original_filename).suffix.lower()
    return upload_root / f"{dataset_id}{suffix}"


def inspect_dataset(file_name: str, content: bytes) -> DatasetInspection:
    """Inspect an uploaded tabular dataset and return row/column metadata."""
    suffix = Path(file_name).suffix.lower()

    if suffix == ".csv":
        frame = pl.read_csv(BytesIO(content))
        return _inspection_from_frame(frame, "csv")
    if suffix == ".json":
        frame = _frame_from_json_bytes(content)
        return _inspection_from_frame(frame, "json")
    if suffix == ".parquet":
        frame = pl.read_parquet(BytesIO(content))
        return _inspection_from_frame(frame, "parquet")
    if suffix == ".xlsx":
        frame = _frame_from_xlsx_bytes(content)
        return _inspection_from_frame(frame, "xlsx")

    raise ValueError(f"Unsupported dataset file type: {suffix}")


def load_dataset_frame(file_path: Path, file_type: str | None = None) -> pl.DataFrame:
    """Load a stored dataset from disk into a Polars DataFrame."""
    normalized_type = (file_type or file_path.suffix.lstrip(".")).lower()

    if normalized_type == "csv":
        return pl.read_csv(file_path)
    if normalized_type == "json":
        return _frame_from_json_bytes(file_path.read_bytes())
    if normalized_type == "parquet":
        return pl.read_parquet(file_path)
    if normalized_type == "xlsx":
        return _frame_from_xlsx_bytes(file_path.read_bytes())

    raise ValueError(f"Unsupported dataset file type: {normalized_type}")


def _inspection_from_frame(frame: pl.DataFrame, file_type: str) -> DatasetInspection:
    return DatasetInspection(
        row_count=frame.height,
        column_count=frame.width,
        columns=[str(column) for column in frame.columns],
        file_type=file_type,
    )


def _frame_from_json_bytes(content: bytes) -> pl.DataFrame:
    payload = json.loads(content.decode("utf-8"))
    return _frame_from_json_payload(payload)


def _frame_from_json_payload(payload: Any) -> pl.DataFrame:
    if isinstance(payload, list):
        if not payload:
            return pl.DataFrame()
        if all(isinstance(item, dict) for item in payload):
            return pl.DataFrame(payload)
        return pl.DataFrame({"value": payload})

    if isinstance(payload, dict):
        if payload and all(isinstance(value, list) for value in payload.values()):
            return pl.DataFrame(payload)
        return pl.DataFrame([payload])

    return pl.DataFrame({"value": [payload]})


def _frame_from_xlsx_bytes(content: bytes) -> pl.DataFrame:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)

        if header_row is None:
            return pl.DataFrame()

        header_cells = _trim_trailing_empty_cells(list(header_row))
        columns = [
            str(value) if value is not None else f"column_{index + 1}"
            for index, value in enumerate(header_cells)
        ]

        normalized_rows: list[dict[str, Any]] = []

        for row in rows:
            cells = _trim_trailing_empty_cells(list(row))
            if not cells:
                continue
            if len(cells) > len(columns):
                for index in range(len(columns), len(cells)):
                    columns.append(f"column_{index + 1}")
            normalized_rows.append(
                {
                    columns[index]: cells[index] if index < len(cells) else None
                    for index in range(len(columns))
                }
            )

        if not normalized_rows:
            return pl.DataFrame(schema=columns)

        return pl.DataFrame(normalized_rows)
    finally:
        workbook.close()


def _trim_trailing_empty_cells(values: list[Any]) -> list[Any]:
    trimmed = list(values)
    while trimmed and trimmed[-1] in (None, ""):
        trimmed.pop()
    return trimmed
